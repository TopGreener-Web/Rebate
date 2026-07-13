from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
WEEK_HEADER_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})-(\d{1,2})\.(\d{1,2})$")
UNTIL_RE = re.compile(r"Until\s+(\d{4})\.(\d{1,2})\.(\d{1,2})", re.IGNORECASE)


def clean_text(value) -> str:
    return "" if value is None else str(value).strip()


def number(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def rounded(value: float) -> float:
    return round(float(value), 2)


def source_name(path: Path) -> str:
    return path.name.split("每周", 1)[0].strip() or path.stem


def canonical_sku(raw_sku: str) -> str:
    sku = clean_text(raw_sku)
    sku = re.sub(r"\s+UPC$", "", sku, flags=re.IGNORECASE)
    sku = re.sub(r"-UPC$", "", sku, flags=re.IGNORECASE)
    return sku.strip()


def product_model(sku: str) -> str:
    value = canonical_sku(sku)
    parts = [part for part in value.split("-") if part]
    if len(parts) >= 2 and parts[0].upper() == "T":
        return f"{parts[0]}-{parts[1]}"
    return parts[0] if parts else value


def extract_pack_size(sku: str) -> int:
    value = canonical_sku(sku).upper()
    matches = re.findall(r"(?:^|[-_\s(])(\d{1,3})\s*(?:PCS?|PACK)(?=$|[-_\s()])", value)
    matches.extend(re.findall(r"(?:^|[-_\s(]|[A-Z])(\d{1,3})P(?=$|[-_\s)])", value))
    sizes = [int(match) for match in matches if int(match) > 1]
    return max(sizes) if sizes else 1


def extract_panel_finish(sku: str) -> str:
    value = f"-{canonical_sku(sku).upper()}-"
    if "-PC-" in value:
        return "PC"
    if "-PB-" in value:
        return "PB"
    return ""


def extract_color(sku: str, model: str) -> str:
    value = canonical_sku(sku)
    remainder = value[len(model) :].lstrip("-") if value.upper().startswith(model.upper()) else value
    if not remainder:
        return "Unknown"
    for token in re.split(r"[-_\s()]+", remainder):
        token = token.strip()
        if not token:
            continue
        if token.upper() in {"UPC", "FNSKU", "STICKER", "STICKERED"}:
            continue
        if re.fullmatch(r"\d+\s*(?:PCS?|PACK|P)", token, flags=re.IGNORECASE):
            continue
        cleaned = re.sub(r"\d+\s*(?:PCS?|PACK|P)$", "", token, flags=re.IGNORECASE).strip()
        if cleaned:
            return cleaned
    return "Unknown"


def header_index(headers: list[str], label: str) -> int | None:
    label_fold = label.casefold()
    for index, header in enumerate(headers):
        if header.casefold() == label_fold:
            return index
    return None


def parse_until_date(path: Path) -> date:
    match = UNTIL_RE.search(path.name)
    if not match:
        raise ValueError(f"Cannot find Until yyyy.mm.dd date in {path.name}")
    year, month, day = [int(part) for part in match.groups()]
    return date(year, month, day)


def find_week_pairs(headers: list[str]) -> list[dict]:
    pairs = []
    index = 0
    while index < len(headers) - 1:
        header = headers[index]
        match = WEEK_HEADER_RE.match(header)
        next_match = WEEK_HEADER_RE.match(headers[index + 1])
        if match and next_match:
            pairs.append(
                {
                    "header": header,
                    "revenue_index": index,
                    "units_index": index + 1,
                    "start_month": int(match.group(1)),
                    "start_day": int(match.group(2)),
                    "end_month": int(match.group(3)),
                    "end_day": int(match.group(4)),
                },
            )
            index += 2
        else:
            index += 1
    return pairs


def assign_week_dates(pairs: list[dict], until_date: date) -> list[dict]:
    if not pairs:
        return []

    current_year = until_date.year
    previous_end_month = pairs[-1]["end_month"]
    dated_reversed = []

    for pair in reversed(pairs):
        end_month = pair["end_month"]
        if end_month > previous_end_month:
            current_year -= 1
        end_year = current_year
        start_year = end_year - 1 if pair["start_month"] > pair["end_month"] else end_year
        dated_reversed.append(
            {
                **pair,
                "start": date(start_year, pair["start_month"], pair["start_day"]),
                "end": date(end_year, pair["end_month"], pair["end_day"]),
                "year": end_year,
                "month": end_month,
            },
        )
        previous_end_month = end_month

    return list(reversed(dated_reversed))


def empty_annual() -> dict:
    return {
        "units": [0.0] * 12,
        "revenue": [0.0] * 12,
    }


def add_month(annual: dict, year: int, month: int, units: float, revenue: float) -> None:
    year_key = str(year)
    if year_key not in annual:
        annual[year_key] = empty_annual()
    annual[year_key]["units"][month - 1] += units
    annual[year_key]["revenue"][month - 1] += revenue


def month_label(months: list[int]) -> str:
    if not months:
        return "No months"
    ranges = []
    start = previous = months[0]
    for month in months[1:]:
        if month == previous + 1:
            previous = month
            continue
        ranges.append((start, previous))
        start = previous = month
    ranges.append((start, previous))

    labels = []
    for start, end in ranges:
        if start == end:
            labels.append(MONTH_LABELS[start - 1])
        else:
            labels.append(f"{MONTH_LABELS[start - 1]}-{MONTH_LABELS[end - 1]}")
    return ", ".join(labels)


def annual_totals(annual: dict) -> dict:
    result = {}
    for year, values in annual.items():
        units = [rounded(value) for value in values["units"]]
        revenue = [rounded(value) for value in values["revenue"]]
        result[year] = {
            "units": units,
            "revenue": revenue,
            "totalUnits": rounded(sum(units)),
            "totalRevenue": rounded(sum(revenue)),
            "unitSource": "Weekly item units (package units x packSize)",
        }
    return result


def product_template(sku: str) -> dict:
    model = product_model(sku)
    return {
        "sku": sku,
        "asinValues": set(),
        "title": "",
        "category": "",
        "model": model,
        "qty": "",
        "color": extract_color(sku, model),
        "status": "Active",
        "packSize": extract_pack_size(sku),
        "panelFinish": extract_panel_finish(sku),
        "annual": {},
        "sourceBreakdown": {},
        "sourceNames": set(),
        "rowChannels": set(),
    }


def build_dashboard_data(source_dir: Path) -> dict:
    products: dict[str, dict] = {}
    source_files = []
    coverage: dict[int, set[int]] = defaultdict(set)
    model_categories: dict[str, Counter] = defaultdict(Counter)

    for workbook_path in sorted(source_dir.glob("*.xlsx")):
        channel = source_name(workbook_path)
        until_date = parse_until_date(workbook_path)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = workbook["Total"] if "Total" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(rows)]
        week_pairs = assign_week_dates(find_week_pairs(headers), until_date)

        asin_index = header_index(headers, "ASIN")
        sku_index = header_index(headers, "SKU")
        style_index = header_index(headers, "Style")
        category_index = header_index(headers, "Category")
        row_channel_index = header_index(headers, "Channel")

        if sku_index is None:
            raise ValueError(f"No SKU column found in {workbook_path.name}")

        source_totals = {"units": 0.0, "revenue": 0.0}
        source_skus: set[str] = set()
        source_years: dict[int, set[int]] = defaultdict(set)
        row_count = 0

        for week in week_pairs:
            coverage[week["year"]].add(week["month"])
            source_years[week["year"]].add(week["month"])

        for row in rows:
            raw_sku = clean_text(row[sku_index] if sku_index < len(row) else "")
            sku = canonical_sku(raw_sku)
            if not sku:
                continue

            row_count += 1
            source_skus.add(sku)
            product = products.setdefault(sku, product_template(sku))
            product["sourceNames"].add(channel)

            asin = clean_text(row[asin_index] if asin_index is not None and asin_index < len(row) else "")
            if asin:
                product["asinValues"].add(asin)

            style = clean_text(row[style_index] if style_index is not None and style_index < len(row) else "")
            if style and not product["qty"]:
                product["qty"] = style

            category = clean_text(row[category_index] if category_index is not None and category_index < len(row) else "")
            if category:
                product["category"] = product["category"] or category
                model_categories[product["model"]][category] += 1

            row_channel = clean_text(row[row_channel_index] if row_channel_index is not None and row_channel_index < len(row) else "")
            if row_channel:
                product["rowChannels"].add(row_channel)

            source_record = product["sourceBreakdown"].setdefault(
                channel,
                {
                    "name": channel,
                    "fileName": workbook_path.name,
                    "annual": {},
                },
            )

            for week in week_pairs:
                revenue = number(row[week["revenue_index"]] if week["revenue_index"] < len(row) else 0)
                package_units = number(row[week["units_index"]] if week["units_index"] < len(row) else 0)
                units = package_units * product["packSize"]
                if not revenue and not units:
                    continue
                add_month(product["annual"], week["year"], week["month"], units, revenue)
                add_month(source_record["annual"], week["year"], week["month"], units, revenue)
                source_totals["units"] += units
                source_totals["revenue"] += revenue

        first_week = week_pairs[0] if week_pairs else None
        last_week = week_pairs[-1] if week_pairs else None
        source_files.append(
            {
                "name": channel,
                "fileName": workbook_path.name,
                "sheet": worksheet.title,
                "firstWeek": f"{first_week['start'].isoformat()} to {first_week['end'].isoformat()}" if first_week else "",
                "lastWeek": f"{last_week['start'].isoformat()} to {last_week['end'].isoformat()}" if last_week else "",
                "weeks": len(week_pairs),
                "rows": row_count,
                "uniqueSkus": len(source_skus),
                "totalUnits": rounded(source_totals["units"]),
                "totalRevenue": rounded(source_totals["revenue"]),
                "coverage": {
                    str(year): {
                        "availableMonths": sorted(months),
                        "coverageLabel": month_label(sorted(months)),
                    }
                    for year, months in sorted(source_years.items())
                },
            },
        )

    final_products = []
    for product in products.values():
        if not product["category"]:
            category_counts = model_categories.get(product["model"])
            product["category"] = category_counts.most_common(1)[0][0] if category_counts else "Uncategorized"

        source_breakdown = []
        for source_record in product["sourceBreakdown"].values():
            annual = annual_totals(source_record["annual"])
            source_breakdown.append(
                {
                    "name": source_record["name"],
                    "fileName": source_record["fileName"],
                    "annual": annual,
                    "totalUnits": rounded(sum(year["totalUnits"] for year in annual.values())),
                    "totalRevenue": rounded(sum(year["totalRevenue"] for year in annual.values())),
                },
            )

        final_products.append(
            {
                "sku": product["sku"],
                "asin": ", ".join(sorted(product["asinValues"])),
                "title": product["title"],
                "category": product["category"],
                "model": product["model"],
                "qty": product["qty"],
                "color": product["color"],
                "status": product["status"],
                "packSize": product["packSize"],
                "panelFinish": product["panelFinish"],
                "channels": sorted(product["sourceNames"]),
                "source": ", ".join(sorted(product["sourceNames"])),
                "rowChannels": sorted(product["rowChannels"]),
                "annual": annual_totals(product["annual"]),
                "sourceBreakdown": sorted(source_breakdown, key=lambda item: item["name"]),
            },
        )

    final_products.sort(key=lambda item: item["sku"])
    years = sorted(coverage)
    year_summaries = {}

    for year in years:
        months = sorted(coverage[year])
        monthly_sales = []
        for month in months:
            units = sum(product["annual"].get(str(year), {}).get("units", [0] * 12)[month - 1] for product in final_products)
            revenue = sum(product["annual"].get(str(year), {}).get("revenue", [0] * 12)[month - 1] for product in final_products)
            monthly_sales.append(
                {
                    "month": month,
                    "label": MONTH_LABELS[month - 1],
                    "units": rounded(units),
                    "revenue": rounded(revenue),
                },
            )

        year_summaries[str(year)] = {
            "availableMonths": months,
            "coverageLabel": month_label(months),
            "monthlySales": monthly_sales,
            "totalUnits": rounded(sum(row["units"] for row in monthly_sales)),
            "totalRevenue": rounded(sum(row["revenue"] for row in monthly_sales)),
            "activeSkus": sum(
                1
                for product in final_products
                if sum(product["annual"].get(str(year), {}).get("units", [0] * 12)[month - 1] for month in months) > 0
            ),
        }

    return {
        "generatedAt": date.today().isoformat(),
        "years": years,
        "defaultYear": 2026 if 2026 in years else years[-1],
        "categories": sorted({product["category"] for product in final_products}),
        "models": sorted({product["model"] for product in final_products}),
        "monthLabels": MONTH_LABELS,
        "yearSummaries": year_summaries,
        "products": final_products,
        "sourceFiles": sorted(source_files, key=lambda item: item["name"]),
    }


def main() -> None:
    project_root = Path(__file__).resolve().parents[1]
    source_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\stephen.deng\Documents\销售数据2")
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else project_root / "src" / "data" / "liderDashboardData.json"
    data = build_dashboard_data(source_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(data['products'])} products from {len(data['sourceFiles'])} sources to {output_path}")


if __name__ == "__main__":
    main()
