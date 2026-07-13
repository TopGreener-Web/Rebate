from __future__ import annotations

import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
PERIODS = {
    "2024": [(2024, month) for month in range(1, 13)],
    "2025": [(2025, month) for month in range(1, 13)],
    "2026_1-5": [(2026, month) for month in range(1, 6)],
}
WINDOW_MONTHS = [month for months in PERIODS.values() for month in months]
MIN_SELLING_MONTHS = 18
LOW_MONTHLY_THRESHOLD = 100


def normalized_base_sku(sku: str) -> str:
    return re.sub(r"(\d{1,2})P$", "", str(sku).strip(), flags=re.IGNORECASE)


def clean_text(value) -> str:
    return "" if value is None else str(value).strip()


def round_one(value: float) -> float:
    return round(float(value), 1)


def is_discontinued(status_values: set[str]) -> bool:
    return any("discontinue" in status.casefold() for status in status_values)


def period_stats(monthly: dict[tuple[int, int], float], months: list[tuple[int, int]]) -> dict:
    values = [float(monthly.get(month, 0)) for month in months]
    selling_values = [value for value in values if value > 0]
    total = sum(values)
    selling_months = len(selling_values)
    return {
        "total": total,
        "selling_months": selling_months,
        "excluded_zero_months": len(values) - selling_months,
        "avg": total / selling_months if selling_months else 0,
        "calendar_avg": total / len(values) if values else 0,
        "peak": max(values) if values else 0,
    }


def build_groups(products: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict] = {}

    for product in products:
        sku = clean_text(product.get("sku"))
        color = clean_text(product.get("color")) or "Unknown"
        key = (normalized_base_sku(sku), color)
        group = grouped.setdefault(
            key,
            {
                "sku_base": key[0],
                "category": clean_text(product.get("category")),
                "model": clean_text(product.get("model")),
                "color": color,
                "statuses": set(),
                "skus": set(),
                "pack_sizes": set(),
                "monthly": defaultdict(float),
            },
        )

        if clean_text(product.get("category")) and not group["category"]:
            group["category"] = clean_text(product.get("category"))
        if clean_text(product.get("model")) and not group["model"]:
            group["model"] = clean_text(product.get("model"))

        group["statuses"].add(clean_text(product.get("status")) or "Blank")
        group["skus"].add(sku)
        pack_size = int(product.get("packSize") or 1)
        group["pack_sizes"].add(pack_size)

        for year, annual in product.get("annual", {}).items():
            year_int = int(year)
            for month_index, units in enumerate(annual.get("units", []), start=1):
                group["monthly"][(year_int, month_index)] += float(units or 0) * pack_size

    return list(grouped.values())


def build_records(groups: list[dict]) -> tuple[list[dict], list[dict]]:
    all_records = []
    for group in groups:
        if is_discontinued(group["statuses"]):
            continue

        stats = {label: period_stats(group["monthly"], months) for label, months in PERIODS.items()}
        window_stats = period_stats(group["monthly"], WINDOW_MONTHS)
        record = {
            **group,
            "stats": stats,
            "window_stats": window_stats,
        }
        all_records.append(record)

    by_model: dict[str, list[dict]] = defaultdict(list)
    for record in all_records:
        if record["window_stats"]["selling_months"] > 0:
            by_model[record["model"]].append(record)

    candidates = []
    for record in all_records:
        if record["window_stats"]["selling_months"] < MIN_SELLING_MONTHS:
            continue
        if not all(record["stats"][period]["avg"] < LOW_MONTHLY_THRESHOLD for period in PERIODS):
            continue

        peers = [
            peer
            for peer in by_model.get(record["model"], [])
            if peer["sku_base"] != record["sku_base"] and peer["color"] != record["color"]
        ]
        if not peers:
            peers = [peer for peer in by_model.get(record["model"], []) if peer["sku_base"] != record["sku_base"]]
        best_peer = max(peers, key=lambda item: item["window_stats"]["avg"], default=None)
        candidates.append({**record, "best_peer": best_peer})

    candidates.sort(key=lambda item: (item["window_stats"]["avg"], item["model"], item["color"], item["sku_base"]))
    return candidates, all_records


def format_status(statuses: set[str]) -> str:
    return ", ".join(sorted(statuses))


def format_skus(skus: set[str]) -> str:
    return ", ".join(sorted(skus))


def main_rows(candidates: list[dict]) -> list[dict]:
    rows = []
    for index, candidate in enumerate(candidates, start=1):
        best_peer = candidate["best_peer"]
        row = {
            "排名": index,
            "SKU组": candidate["sku_base"],
            "包含SKU": format_skus(candidate["skus"]),
            "类别": candidate["category"],
            "型号": candidate["model"],
            "颜色": candidate["color"],
            "状态": format_status(candidate["statuses"]),
            "PackSize": ", ".join(str(size) for size in sorted(candidate["pack_sizes"])),
            "2024月销_单只": round_one(candidate["stats"]["2024"]["avg"]),
            "2025月销_单只": round_one(candidate["stats"]["2025"]["avg"]),
            "2026_1-5月销_单只": round_one(candidate["stats"]["2026_1-5"]["avg"]),
            "全期月销_单只": round_one(candidate["window_stats"]["avg"]),
            "2024合计_单只": round_one(candidate["stats"]["2024"]["total"]),
            "2025合计_单只": round_one(candidate["stats"]["2025"]["total"]),
            "2026_1-5合计_单只": round_one(candidate["stats"]["2026_1-5"]["total"]),
            "有销售月份_29": candidate["window_stats"]["selling_months"],
            "排除0销量月份": candidate["window_stats"]["excluded_zero_months"],
            "最高单月销量_单只": round_one(candidate["window_stats"]["peak"]),
            "同型号最佳颜色": best_peer["color"] if best_peer else "",
            "同型号最佳SKU组": best_peer["sku_base"] if best_peer else "",
            "同型号最佳全期月销_单只": round_one(best_peer["window_stats"]["avg"]) if best_peer else "",
            "同型号最佳2024月销_单只": round_one(best_peer["stats"]["2024"]["avg"]) if best_peer else "",
            "同型号最佳2025月销_单只": round_one(best_peer["stats"]["2025"]["avg"]) if best_peer else "",
            "同型号最佳2026_1-5月销_单只": round_one(best_peer["stats"]["2026_1-5"]["avg"]) if best_peer else "",
        }
        row["对比说明"] = (
            f"候选 {candidate['color']} 全期月销 {row['全期月销_单只']}；"
            f"同型号最佳 {row['同型号最佳颜色']} 月销 {row['同型号最佳全期月销_单只']}"
            if best_peer
            else "同型号没有其它颜色/版本可对比"
        )
        rows.append(row)
    return rows


def monthly_rows(candidates: list[dict]) -> list[dict]:
    rows = []
    for candidate in candidates:
        row = {
            "SKU组": candidate["sku_base"],
            "型号": candidate["model"],
            "颜色": candidate["color"],
            "包含SKU": format_skus(candidate["skus"]),
        }
        for year, month in WINDOW_MONTHS:
            label = f"{year}-{MONTH_LABELS[month - 1]}"
            row[label] = round_one(candidate["monthly"].get((year, month), 0))
        rows.append(row)
    return rows


def write_csv(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def append_sheet(workbook: Workbook, title: str, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["No rows"])
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="DCE9E5")
    header_font = Font(bold=True, color="1D2523")
    thin = Side(style="thin", color="DCE3DF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        header = str(column_cells[0].value or "")
        width = min(max(len(header) + 2, 12), 34)
        if header in {"包含SKU", "对比说明"}:
            width = 52
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_workbook(path: Path, main: list[dict], detail: list[dict], candidate_count: int, all_count: int) -> None:
    workbook = Workbook()
    method = workbook.active
    method.title = "Method"
    method_rows = [
        ["品牌", "Lider"],
        ["筛选范围", "2024 Jan-Dec, 2025 Jan-Dec, 2026 Jan-May"],
        ["入选状态", "未标注为 discontinue；当前数据中 Active/Inactive/Blank 均保留"],
        ["月销算法", "单只销量合计 / 有销售月份数；0 销售月份不进分母，用作断货/无货段排除"],
        ["多包装算法", "SKU 销量 * packSize 后折算为单只；同一 SKU 组的 pack 版本合并"],
        ["颜色算法", "不同颜色不合并；同型号最佳颜色列用于对比"],
        ["最短销售历史", f"有销售月份少于 {MIN_SELLING_MONTHS} 的 SKU 组先不考虑"],
        ["低销阈值", f"2024、2025、2026 Jan-May 的有效月销均小于 {LOW_MONTHLY_THRESHOLD} 单只"],
        ["候选数", candidate_count],
        ["可评估SKU组数", all_count],
    ]
    for row in method_rows:
        method.append(row)
    method.column_dimensions["A"].width = 22
    method.column_dimensions["B"].width = 92
    for row in method.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].font = Font(bold=True)

    append_sheet(workbook, "Discontinue Candidates", main)
    append_sheet(workbook, "Monthly Detail", detail)
    workbook.save(path)


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: build_discontinue_candidates.py <liderDashboardData.json> <output-dir>", file=sys.stderr)
        return 2

    input_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)

    data = json.loads(input_path.read_text(encoding="utf-8"))
    groups = build_groups(data["products"])
    candidates, all_records = build_records(groups)
    main = main_rows(candidates)
    detail = monthly_rows(candidates)

    csv_path = output_dir / "discontinue_candidates_2026.csv"
    detail_csv_path = output_dir / "discontinue_candidates_2026_monthly_detail.csv"
    workbook_path = output_dir / "discontinue_candidates_2026.xlsx"

    write_csv(csv_path, main)
    write_csv(detail_csv_path, detail)
    write_workbook(workbook_path, main, detail, len(candidates), len(all_records))

    print(
        json.dumps(
            {
                "candidate_count": len(candidates),
                "evaluated_groups": len(all_records),
                "xlsx": str(workbook_path),
                "csv": str(csv_path),
                "monthly_csv": str(detail_csv_path),
            },
            ensure_ascii=False,
            indent=2,
        ),
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
