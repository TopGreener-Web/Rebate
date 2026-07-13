from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("C:/Users/stephen.deng/Documents/\u9500\u552e\u6570\u636e/\u9500\u552e\u6570\u636e/Lider")

MONTH_ALIASES = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
    "december": 12,
}

MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
TRACKING_HEADER_RE = re.compile(
    r"(?P<year>20\d{2})\s+(?P<month>Jan|Feb|Mar|Apr|April|May|Jun|June|Jul|July|Aug|Sep|Sept|Oct|Nov|Dec|December)\s+Seller",
    re.IGNORECASE,
)
WEEK_HEADER_RE = re.compile(r"^(?P<sm>\d{1,2})\.(?P<sd>\d{1,2})-(?P<em>\d{1,2})\.(?P<ed>\d{1,2})$")
REVENUE_HEADER_RE = re.compile(r"^(?P<year>20\d{2})\s+Sales Revenue$", re.IGNORECASE)


def to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.upper() == "NULL":
        return 0.0
    text = text.replace("$", "").replace(",", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def model_key(sku: str) -> str:
    parts = [part for part in sku.split("-") if part]
    if len(parts) >= 2:
        return f"{parts[0]}-{parts[1]}"
    return sku or "Uncategorized"


def pack_size(sku: str) -> int:
    match = re.search(r"(\d{1,2})P$", clean_text(sku).upper())
    return int(match.group(1)) if match else 1


def finish_code(sku: str) -> str:
    parts = [part for part in clean_text(sku).upper().split("-") if part]
    while len(parts) > 1 and parts[-1] in {"CA", "WP", "NEW"}:
        parts.pop()
    if not parts:
        return ""
    code = parts[-1]
    code = re.sub(r"(\d{1,2})P$", "", code)
    code = re.sub(r"WP$", "", code)
    return code


def panel_finish(sku: str) -> str:
    code = finish_code(sku)
    return code if code in {"PC", "PB"} else ""


def normalize_header(value) -> str:
    return clean_text(value).lower()


def make_annual_record() -> dict:
    return {
        "mainUnits": [0.0] * 12,
        "weeklyUnits": [0.0] * 12,
        "revenue": [0.0] * 12,
        "annualRevenue": 0.0,
        "mainAvailable": False,
        "weeklyAvailable": False,
    }


products: dict[str, dict] = {}
year_months_from_main: dict[int, set[int]] = defaultdict(set)
year_months_from_weekly: dict[int, set[int]] = defaultdict(set)
source_files: list[dict] = []


def ensure_product(sku: str) -> dict:
    sku = clean_text(sku)
    product = products.get(sku)
    if product is None:
        product = {
            "sku": sku,
            "asin": "",
            "title": "",
            "category": "",
            "model": "",
            "qty": "",
            "color": "",
            "status": "",
            "annual": defaultdict(make_annual_record),
        }
        products[sku] = product
    return product


def update_metadata(product: dict, values: dict[str, str]) -> None:
    for key, value in values.items():
        text = clean_text(value)
        if text and not product.get(key):
            product[key] = text


def worksheet_by_name(workbook, desired_name: str):
    desired = desired_name.casefold()
    for sheet_name in workbook.sheetnames:
        if sheet_name.casefold() == desired:
            return workbook[sheet_name]
    return None


def parse_tracking_workbook(path: Path, year: int) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = worksheet_by_name(workbook, "Main")
        if sheet is None:
            return

        rows = sheet.iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(rows, [])]
        header_lookup = {normalize_header(value): index for index, value in enumerate(headers)}

        sku_col = header_lookup.get("sku", header_lookup.get("seller-sku"))
        asin_col = header_lookup.get("asin")
        if sku_col is None:
            return

        month_cols: list[tuple[int, int]] = []
        for index, header in enumerate(headers):
            match = TRACKING_HEADER_RE.search(header)
            if not match:
                continue
            header_year = int(match.group("year"))
            if header_year != year:
                continue
            month = MONTH_ALIASES[match.group("month").lower()]
            month_cols.append((index, month))
            year_months_from_main[year].add(month)

        if not month_cols:
            return

        for row in rows:
            sku = clean_text(row[sku_col] if sku_col < len(row) else "")
            if not sku:
                continue

            product = ensure_product(sku)
            asin = clean_text(row[asin_col] if asin_col is not None and asin_col < len(row) else "")
            update_metadata(product, {"asin": asin})
            annual = product["annual"][year]
            annual["mainAvailable"] = True

            for column_index, month in month_cols:
                value = row[column_index] if column_index < len(row) else None
                annual["mainUnits"][month - 1] += to_number(value)
    finally:
        workbook.close()


def infer_week_pairs(headers: list[str]) -> list[dict]:
    pairs = []
    current_year = 2022
    previous_start: tuple[int, int] | None = None
    index = 0

    while index < len(headers) - 1:
        header = clean_text(headers[index])
        if header and header == clean_text(headers[index + 1]):
            match = WEEK_HEADER_RE.match(header)
            if match:
                start_month = int(match.group("sm"))
                start_day = int(match.group("sd"))
                end_month = int(match.group("em"))
                end_day = int(match.group("ed"))
                start_key = (start_month, start_day)

                if previous_start is not None and start_key < previous_start:
                    current_year += 1

                end_year = current_year + (1 if end_month < start_month else 0)
                pairs.append(
                    {
                        "label": header,
                        "revenue_col": index,
                        "units_col": index + 1,
                        "end_year": end_year,
                        "end_month": end_month,
                        "end_day": end_day,
                    }
                )
                previous_start = start_key
                index += 2
                continue
        index += 1

    return pairs


def parse_weekly_workbook(path: Path, year: int) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = worksheet_by_name(workbook, "Total")
        if sheet is None:
            return

        rows = sheet.iter_rows(values_only=True)
        raw_headers = list(next(rows, []))
        headers = [clean_text(value) for value in raw_headers]
        header_lookup = {normalize_header(value): index for index, value in enumerate(headers) if value}

        sku_col = header_lookup.get("sku", header_lookup.get("seller-sku"))
        if sku_col is None:
            return

        revenue_col = None
        for index, header in enumerate(headers):
            match = REVENUE_HEADER_RE.match(header)
            if match and int(match.group("year")) == year:
                revenue_col = index
                break

        field_columns = {
            "asin": header_lookup.get("asin"),
            "title": header_lookup.get("title"),
            "category": header_lookup.get("category"),
            "model": header_lookup.get("style"),
            "qty": header_lookup.get("qty"),
            "color": header_lookup.get("color"),
            "status": header_lookup.get("status"),
        }
        pairs = [pair for pair in infer_week_pairs(headers) if pair["end_year"] == year]
        for pair in pairs:
            year_months_from_weekly[year].add(pair["end_month"])

        for row in rows:
            sku = clean_text(row[sku_col] if sku_col < len(row) else "")
            if not sku:
                continue

            product = ensure_product(sku)
            metadata = {}
            for field, column in field_columns.items():
                if column is not None and column < len(row):
                    metadata[field] = row[column]
            update_metadata(product, metadata)

            annual = product["annual"][year]
            annual["weeklyAvailable"] = True
            if revenue_col is not None and revenue_col < len(row):
                annual["annualRevenue"] += to_number(row[revenue_col])

            for pair in pairs:
                month_index = pair["end_month"] - 1
                revenue_value = row[pair["revenue_col"]] if pair["revenue_col"] < len(row) else None
                units_value = row[pair["units_col"]] if pair["units_col"] < len(row) else None
                annual["revenue"][month_index] += to_number(revenue_value)
                annual["weeklyUnits"][month_index] += to_number(units_value)
    finally:
        workbook.close()


def workbook_year(path: Path) -> int | None:
    try:
        return int(path.parent.name)
    except ValueError:
        match = re.search(r"(20\d{2})", path.name)
        return int(match.group(1)) if match else None


def iter_source_workbooks() -> list[Path]:
    ignored_fragments = ["\u52a0\u62ff\u5927", "\u82f1\u56fd"]
    paths = []
    for path in ROOT.rglob("*.xlsx"):
        if path.name.startswith("~$"):
            continue
        if any(fragment in path.name for fragment in ignored_fragments):
            continue
        paths.append(path)
    return sorted(paths, key=lambda item: (item.parent.name, item.name))


def round_value(value: float, digits: int = 2) -> float:
    rounded = round(float(value), digits)
    return 0.0 if rounded == -0.0 else rounded


def build_payload() -> dict:
    years = sorted(
        {
            int(year)
            for product in products.values()
            for year, annual in product["annual"].items()
            if annual["mainAvailable"] or annual["weeklyAvailable"]
        }
    )

    serialized_products = []
    for product in products.values():
        annual_payload = {}
        for year in years:
            annual = product["annual"].get(year)
            if not annual:
                continue

            use_main = bool(annual["mainAvailable"])
            units = annual["mainUnits"] if use_main else annual["weeklyUnits"]
            revenue = annual["revenue"]
            total_revenue = sum(revenue)
            total_units = sum(units)

            if not any(units) and not total_revenue:
                continue

            annual_payload[str(year)] = {
                "units": [round_value(value, 0) for value in units],
                "revenue": [round_value(value) for value in revenue],
                "totalUnits": round_value(total_units, 0),
                "totalRevenue": round_value(total_revenue),
                "unitSource": "Main" if use_main else "Total",
            }

        if not annual_payload:
            continue

        serialized_products.append(
            {
                "sku": product["sku"],
                "asin": product["asin"],
                "title": product["title"],
                "category": product["category"] or "Uncategorized",
                "model": product["model"] or model_key(product["sku"]),
                "qty": product["qty"],
                "color": product["color"],
                "status": product["status"],
                "packSize": pack_size(product["sku"]),
                "panelFinish": panel_finish(product["sku"]),
                "annual": annual_payload,
            }
        )

    serialized_products.sort(key=lambda item: (item["category"], item["model"], item["sku"]))

    categories = sorted({product["category"] for product in serialized_products if product["category"]})
    models = sorted({product["model"] for product in serialized_products if product["model"]})
    year_summaries = {}
    for year in years:
        months = sorted(year_months_from_main.get(year) or year_months_from_weekly.get(year) or [])
        monthly_sales = []
        for month in months:
            units = sum(product["annual"].get(str(year), {}).get("units", [0] * 12)[month - 1] for product in serialized_products)
            revenue = sum(
                product["annual"].get(str(year), {}).get("revenue", [0] * 12)[month - 1]
                for product in serialized_products
            )
            monthly_sales.append(
                {
                    "month": month,
                    "label": MONTH_LABELS[month - 1],
                    "units": round_value(units, 0),
                    "revenue": round_value(revenue),
                }
            )

        total_units = sum(product["annual"].get(str(year), {}).get("totalUnits", 0) for product in serialized_products)
        total_revenue = sum(product["annual"].get(str(year), {}).get("totalRevenue", 0) for product in serialized_products)
        active_skus = sum(1 for product in serialized_products if product["annual"].get(str(year), {}).get("totalUnits", 0) > 0)

        year_summaries[str(year)] = {
            "availableMonths": months,
            "coverageLabel": coverage_label(months),
            "monthlySales": monthly_sales,
            "totalUnits": round_value(total_units, 0),
            "totalRevenue": round_value(total_revenue),
            "activeSkus": active_skus,
        }

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "years": years,
        "defaultYear": years[-1] if years else None,
        "categories": categories,
        "models": models,
        "monthLabels": MONTH_LABELS,
        "yearSummaries": year_summaries,
        "products": serialized_products,
        "sourceFiles": source_files,
    }


def coverage_label(months: list[int]) -> str:
    if not months:
        return "No months"
    if months == list(range(1, 13)):
        return "Jan-Dec"
    if len(months) == 1:
        return MONTH_LABELS[months[0] - 1]
    return f"{MONTH_LABELS[months[0] - 1]}-{MONTH_LABELS[months[-1] - 1]}"


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: extract_lider_dashboard_data.py <output-json>", file=sys.stderr)
        return 2

    output_path = Path(sys.argv[1])
    for path in iter_source_workbooks():
        year = workbook_year(path)
        if year is None:
            continue

        if "Sales Tracking" in path.name:
            parse_tracking_workbook(path, year)
            source_type = "tracking-main"
        elif "LIDER\u6bcf\u5468\u9500\u91cf\u6982\u51b5Until" in path.name:
            parse_weekly_workbook(path, year)
            source_type = "weekly-total"
        else:
            continue

        source_files.append(
            {
                "year": year,
                "type": source_type,
                "fileName": path.name,
            }
        )

    payload = build_payload()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"Wrote {len(payload['products'])} products across {len(payload['years'])} years to {output_path}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
